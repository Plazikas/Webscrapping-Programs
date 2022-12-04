from calendar import c
from bs4 import BeautifulSoup
import requests
import openpyxl

temporada = 'temporada_22-23/'
dir_path = '/home/pedro/Escritorio/DatosFutbol/' + temporada 
excel_clasificacion = openpyxl.load_workbook(dir_path + 'clasificacion.xlsx')

for i in range (1,13):
    jornada = 'Jornada' + str(i)
    if jornada not in excel_clasificacion.sheetnames:
        sheet_jornada = excel_clasificacion.create_sheet(jornada)

    sheet_jornada = excel_clasificacion[jornada]

    website = "https://www.resultados-futbol.com/primera/grupo1/jornada" + str(i) 
    agent = {"User-Agent":"Mozilla/5.0"}
    result = requests.get(website, headers=agent)
    content = result.text

    soup = BeautifulSoup(content, 'lxml')
    box_clasificacion = soup.find('div', {'id':'col-clasificacion'})
    body_clasificacion = box_clasificacion.find('tbody')
    equipos = body_clasificacion.find_all('tr')

    datos_equipos = []
    print('jornada' + str(i))

    for equipo in equipos:
        posicion = equipo.find('th').get_text()
        nombre = equipo.find('td', {'class':'equipo'}).find('a').get('href')[1:]
        puntos = equipo.find('td', {'class':'pts'}).get_text()
        jugados = equipo.find('td', {'class':'pj'}).get_text()
        victorias = equipo.find('td', {'class':'win'}).get_text()
        empates = equipo.find('td', {'class':'draw'}).get_text()
        derrotas = equipo.find('td', {'class':'lose'}).get_text()
        goles_favor = equipo.find('td', {'class':'f'}).get_text()
        goles_contra = equipo.find('td', {'class':'c'}).get_text()

        datos_equipo = (posicion, nombre, puntos, jugados, victorias, empates, derrotas, goles_favor, goles_contra)
        datos_equipos.append(datos_equipo)
        
    pos_row = 1
    pos_col = 1
    sheet_jornada.cell(pos_row,pos_col).value = 'Posición'
    sheet_jornada.cell(pos_row, pos_col + 1).value = 'Equipo'
    sheet_jornada.cell(pos_row, pos_col + 2).value = 'Puntos'
    sheet_jornada.cell(pos_row, pos_col + 3).value = 'Jugados'
    sheet_jornada.cell(pos_row, pos_col + 4).value = 'Victorias'
    sheet_jornada.cell(pos_row, pos_col + 5).value = 'Empates'
    sheet_jornada.cell(pos_row, pos_col + 6).value = 'Derrotas'
    sheet_jornada.cell(pos_row, pos_col + 7).value = 'G. Favor'
    sheet_jornada.cell(pos_row, pos_col + 8).value = 'G. Contra'
    sheet_jornada.cell(pos_row, pos_col + 9).value = 'Gol Average'

    for datos_equipo in datos_equipos:
        print(datos_equipo)
        pos_row = pos_row + 1
        sheet_jornada.cell(pos_row,pos_col).value = datos_equipo[0]
        sheet_jornada.cell(pos_row, pos_col + 1).value = datos_equipo[1]
        sheet_jornada.cell(pos_row, pos_col + 2).value = datos_equipo[2]
        sheet_jornada.cell(pos_row, pos_col + 3).value = datos_equipo[3]
        sheet_jornada.cell(pos_row, pos_col + 4).value = datos_equipo[4]
        sheet_jornada.cell(pos_row, pos_col + 5).value = datos_equipo[5]
        sheet_jornada.cell(pos_row, pos_col + 6).value = datos_equipo[6]
        sheet_jornada.cell(pos_row, pos_col + 7).value = datos_equipo[7]
        sheet_jornada.cell(pos_row, pos_col + 8).value = datos_equipo[8]
        if int(datos_equipo[8]) > int(datos_equipo[7]):
            sheet_jornada.cell(pos_row, pos_col + 9).value = '-' + str((int(datos_equipo[7]) - int(datos_equipo[8])))
        else:
            sheet_jornada.cell(pos_row, pos_col + 9).value = '+' + str((int(datos_equipo[7]) - int(datos_equipo[8])))
        
    datos_equipos.clear

excel_clasificacion.save(dir_path + 'clasificacion.xlsx')