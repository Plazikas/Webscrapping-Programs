'''
# Programa que toma las estadísticas del acta de los partidos
'''

from csv import excel
from doctest import ELLIPSIS_MARKER
from os import environ
from unittest import result
from bs4 import BeautifulSoup
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def TipoEvento(evento):
    if evento == 'Gol de ' or evento == 'Gol de penalti' or evento == 'Gol de falta' or evento == 'Gol de (p.p)':
        return 'Goles'
    elif evento == 'T. Amarilla' or evento == 'T. Roja' or evento == '2a Amarilla y Roja' or evento == 'Tarjeta Roja a ':
        return 'Tarjetas'
    elif evento == 'Tiro al palo' or evento == 'Gol anulado' or evento == 'Penalti parado' or evento == 'Penalti fallado' or evento == 'Anulado por var':
        return 'Ocasiones'
    elif evento == 'Entra en el partido' or evento == 'Sale del partido':
        return 'Cambios'
    elif evento == 'Asistencia' or evento == 'Lesionado' or evento == 'Penalti cometido':
        return 'Otros'
    else:
        print(evento)
        return 'Desconocido'

# Abrimos el excel con las jornadas y obtenemos las direcciones de los partidos
temporada = 'temporada_22-23/'
dir_path = '/home/pedro/Escritorio/DatosFutbol/' + temporada

nJornada = 12
excel_jornada = openpyxl.load_workbook(dir_path + 'Partidos/Jornada' + str(nJornada) + '.xlsx') 

partidos_jornada = excel_jornada.sheetnames

for partido in partidos_jornada:
    sheetPartido = excel_jornada[partido]
    dir_partido = partido.split('_')
    equipo_local = dir_partido[0]
    equipo_visitante = dir_partido[1]
    dir_partido = '/partido/' + equipo_local + '/' + equipo_visitante

    print (sheetPartido)

    # Accedemos al acta del partido
    agent = {"User-Agent":"Mozilla/5.0"}
    pagina_partido = 'https://www.resultados-futbol.com/' + dir_partido
    result = requests.get(pagina_partido, headers=agent)
    content = result.content

    soup = BeautifulSoup(content, 'lxml')

    # Obtenemos la fecha del partido 
    box_marcador = soup.find('div', {'id':'marcador'})
    horario = box_marcador.find('span', {'class':'jor-date'}).get_text()
    marcador = box_marcador.find('div', {'class':'resultado resultadoH'}).find_all('span')
    marcador = marcador[0].get_text() + '-' + marcador[1].get_text()

    # Guardamos la fecha, NºJornada, Equipo Local, Equipo Visitante y resultado
    sheetPartido.cell(1,1).value = 'Fecha'
    sheetPartido.cell(1,2).value = horario

    sheetPartido.cell(2,1).value = 'Jornada'
    sheetPartido.cell(2,2).value = nJornada

    sheetPartido.cell(3,1).value = 'Local'
    sheetPartido.cell(3,2).value = equipo_local

    sheetPartido.cell(4,1).value = 'Visitante'
    sheetPartido.cell(4,2).value = equipo_visitante

    sheetPartido.cell(5,1).value = 'Marcador'
    sheetPartido.cell(5,2).value = marcador
 
    # Obtenemos los eventos del partido (Goles, Tarjetas, Cambios, Otros...)
    box_eventos = soup.find('div', {'id':'teams_box'}).find('div', {'class', 'contentitem'})
    nombre_eventos = box_eventos.find_all('h4', {'class':'sepevento'})
    lista_eventos = [['Goles'], ['Tarjetas'], ['Cambios'], ['Ocasiones'], ['Otros']]
    

    eventos = box_eventos.find_all('div',{'class':'evento'})
    for evento in eventos:
        spans = evento.find_all('span')
        datos_evento = []
        for span in spans:
            datos_evento.append(span.get_text())
        tipo_evento = evento.find('small').get_text()
        datos_evento.append(tipo_evento)
        tipo_evento = TipoEvento(tipo_evento)
        
        datos_evento.append(evento.find('a').get_text())
        for i in lista_eventos:
            if i[0] == tipo_evento:
                i.append(datos_evento) 
    diccionario_eventos = {}       
    for i in lista_eventos:
        diccionario_eventos[i[0]] = i[1:]

    # Guardamos los eventos del partido 
        # GUARDAMOS LOS GOLES
    pos_row_goles = 1
    pos_col_goles = 5
    sheetPartido.cell(pos_row_goles, pos_col_goles).value = 'Goles'
    sheetPartido.cell(pos_row_goles, pos_col_goles).font = Font(bold=True)
    sheetPartido.cell(pos_row_goles, pos_col_goles).alignment = Alignment(horizontal='center')
    sheetPartido.merge_cells(start_row=pos_row_goles, start_column=pos_col_goles, end_row=pos_row_goles, end_column=pos_col_goles + 4)
    goles = diccionario_eventos.get('Goles')
    resultado = [0,0]
    try:
        for gol in goles:
            equipo_goleador = ''    # Equipo que marca el gol
            if gol[2] == '':
                equipo_goleador = equipo_local
            else:
                equipo_goleador = equipo_visitante 
            pos_row_goles = pos_row_goles + 1
            sheetPartido.cell(pos_row_goles, pos_col_goles).value = equipo_goleador

            sheetPartido.cell(pos_row_goles, pos_col_goles + 1).value = gol[3]  # Forma en la que se marca el gol

            sheetPartido.cell(pos_row_goles, pos_col_goles + 2 ).value = gol[4] # Jugador que marca el gol

            minuto_gol = gol[1]
            minuto_gol = int(minuto_gol[7:-1])
            sheetPartido.cell(pos_row_goles, pos_col_goles + 3).value = minuto_gol  # Minuto en el que se marca el gol

            if equipo_goleador == equipo_local:
                resultado[0] = resultado[0] + 1
            elif equipo_goleador == equipo_visitante:
                resultado[1] = resultado[1] + 1
            else:
                print('Ha ocurrido un error')
            sheetPartido.cell(pos_row_goles, pos_col_goles + 4).value = str(resultado[0]) + '-' + str(resultado[1]) # Marcador tras el gol marcado
    except:
        print('No hay goles')

        # Guardamos las tarjetas
    pos_row_tarjetas = 1
    pos_col_tarjetas = 11

    sheetPartido.cell(pos_row_tarjetas, pos_col_tarjetas).value = 'Tarjetas'
    sheetPartido.cell(pos_row_tarjetas, pos_col_tarjetas).font = Font(bold=True)
    sheetPartido.cell(pos_row_tarjetas, pos_col_tarjetas).alignment = Alignment(horizontal='center')
    sheetPartido.merge_cells(start_row=pos_row_tarjetas, start_column=pos_col_tarjetas, end_row=pos_row_tarjetas, end_column=pos_col_tarjetas + 3)
    tarjetas = diccionario_eventos.get('Tarjetas')

    try:
        for tarjeta in tarjetas:
            equipo_tarjeta = ''
            if tarjeta[2] == '':
                equipo_tarjeta = equipo_local
            else:
                equipo_tarjeta = equipo_visitante 
            pos_row_tarjetas = pos_row_tarjetas + 1
            sheetPartido.cell(pos_row_tarjetas, pos_col_tarjetas).value = equipo_tarjeta    # Equipo al que le sacan tarjeta

            sheetPartido.cell(pos_row_tarjetas, pos_col_tarjetas + 1).value = tarjeta[3]    # Tipo de tarjeta

            sheetPartido.cell(pos_row_tarjetas, pos_col_tarjetas + 2).value = tarjeta[4]    # Jugador al que le sacan tarjeta

            minuto_tarjeta = tarjeta[1]
            minuto_tarjeta = int(minuto_tarjeta[7:-1])
            sheetPartido.cell(pos_row_tarjetas, pos_col_tarjetas + 3).value = minuto_tarjeta  # Minuto en el que se marca el gol
    except:
        print('No hay tarjetas')

        # Guardamos las ocasiones
    pos_row_ocasiones = 1
    pos_col_ocasiones = 16

    sheetPartido.cell(pos_row_ocasiones, pos_col_ocasiones).value = 'Ocasiones'
    sheetPartido.cell(pos_row_ocasiones, pos_col_ocasiones).font = Font(bold=True)
    sheetPartido.cell(pos_row_ocasiones, pos_col_ocasiones).alignment = Alignment(horizontal='center')
    sheetPartido.merge_cells(start_row=pos_row_ocasiones, start_column=pos_col_ocasiones, end_row=pos_row_ocasiones, end_column=pos_col_ocasiones + 3)
    ocasiones = diccionario_eventos.get('Ocasiones')

    try:
        for ocasion in ocasiones:
            equipo_ocasion = ''
            if ocasion[2] == '':
                equipo_ocasion = equipo_local
            else:
                equipo_ocasion = equipo_visitante 
            pos_row_ocasiones = pos_row_ocasiones + 1
            sheetPartido.cell(pos_row_ocasiones, pos_col_ocasiones).value = equipo_ocasion  # Equipo que provoca la ocasión
            
            sheetPartido.cell(pos_row_ocasiones, pos_col_ocasiones+1).value = ocasion[3]    # Tipo de ocaión

            sheetPartido.cell(pos_row_ocasiones, pos_col_ocasiones + 2).value = ocasion[4]  # Jugador que provoca la ocasión

            minuto_ocasion = ocasion[1]
            minuto_ocasion = int(minuto_ocasion[7:-1])
            sheetPartido.cell(pos_row_ocasiones, pos_col_ocasiones + 3).value = minuto_ocasion  # Minuto en el que se provoca la ocasión

    except:
        print('No hay ocasiones')

        # Guardamos los Otros
    pos_row_otros = 1
    pos_col_otros = 21

    sheetPartido.cell(pos_row_otros, pos_col_otros).value = 'Otros'
    sheetPartido.cell(pos_row_otros, pos_col_otros).font = Font(bold=True)
    sheetPartido.cell(pos_row_otros, pos_col_otros).alignment = Alignment(horizontal='center')
    sheetPartido.merge_cells(start_row=pos_row_otros, start_column=pos_col_otros, end_row=pos_row_otros, end_column=pos_col_otros + 3)
    otros = diccionario_eventos.get('Otros')
    try:
        for otro in otros:
            equipo_otro = ''
            if otro[2] == '':
                equipo_otro = equipo_local
            else:
                equipo_otro = equipo_visitante 
            pos_row_otros = pos_row_otros + 1
            sheetPartido.cell(pos_row_otros, pos_col_otros).value = equipo_otro  # Equipo que provoca otro

            sheetPartido.cell(pos_row_otros, pos_col_otros+1).value = otro[3]    # Tipo de otro

            sheetPartido.cell(pos_row_otros, pos_col_otros + 2).value = otro[4]  # Jugador que provoca otro

            minuto_otro = otro[1]
            minuto_otro = int(minuto_otro[7:-1])
            sheetPartido.cell(pos_row_otros, pos_col_otros + 3).value = minuto_otro  # Minuto en el que se provoca la ocasión

    except:
        print('No hay otros')

        # Guardamos los cambios

    pos_row_cambios = 1
    pos_col_cambios = 26

    sheetPartido.cell(pos_row_cambios, pos_col_cambios).value = 'Cambios'
    sheetPartido.cell(pos_row_cambios, pos_col_cambios).font = Font(bold=True)
    sheetPartido.cell(pos_row_cambios, pos_col_cambios).alignment = Alignment(horizontal='center')
    sheetPartido.merge_cells(start_row=pos_row_cambios, start_column=pos_col_cambios, end_row=pos_row_cambios, end_column=pos_col_cambios + 3)
    cambios = diccionario_eventos.get('Cambios')
    try:
        for cambio in cambios:
            equipo_cambio = ''
            if cambio[2] == '':
                equipo_cambio = equipo_local
            else:
                equipo_cambio = equipo_visitante 
            pos_row_cambios = pos_row_cambios + 1
            sheetPartido.cell(pos_row_cambios, pos_col_cambios).value = equipo_cambio  # Equipo que hace cambio

            sheetPartido.cell(pos_row_cambios, pos_col_cambios+1).value = cambio[3]    # Tipo de cambio

            sheetPartido.cell(pos_row_cambios, pos_col_cambios + 2).value = cambio[4]  # Jugador que sale/entra

            minuto_cambio = cambio[1]
            minuto_cambio = int(minuto_cambio[7:-1])
            sheetPartido.cell(pos_row_cambios, pos_col_cambios + 3).value = minuto_cambio  # Minuto en el que se produce el cambio
    except:
       print('No hay cambios')

    ####
    # Obtenemos las estadísticas del partido
    columna_primera = soup.find('div', {'id':'columna_primera'})
    box_estadistica = columna_primera.find('div',{'id':'box-tabla'})
    estadisticas_class = box_estadistica.find('div', {'class':'contentitem'})
    estadisticas_vec = estadisticas_class.find_all('tr', {'class':'barstyle bar4'})

    estadisticas = []

    for estadistica in estadisticas_vec:
        datos = estadistica.find_all('td')
        estadistica_local = datos[0].get_text()
        estadistica_nombre = datos[1].get_text()
        estadistica_visitante = datos[2].get_text()

        estadisticas.append((estadistica_local, estadistica_nombre, estadistica_visitante))

    # Guardamos los datos de las estadísticas
    pos_row_estadistica = 8
    
    sheetPartido.cell(pos_row_estadistica - 1,1).value = equipo_local
    sheetPartido.cell(pos_row_estadistica - 1,2).value = 'Tipo Estadística'
    sheetPartido.cell(pos_row_estadistica - 1,3).value = equipo_visitante

    for estadistica in estadisticas:
        for j in range (1, len(estadistica)+1):
            sheetPartido.cell(pos_row_estadistica,j).value = estadistica[j-1]
        pos_row_estadistica = pos_row_estadistica + 1

    # Guardamos los cambios realizados
    excel_jornada.save(dir_path + 'Partidos/Jornada' + str(nJornada) + '.xlsx')

    print ('Archivo ' + dir_path + 'Partidos/Jornada' + str(nJornada) + '.xlsx : Actualizado' )